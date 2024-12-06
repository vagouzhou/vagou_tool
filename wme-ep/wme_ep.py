import requests
import pandas as pd
import re
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment

source_file = '/Users/vagouz/Downloads/orignal.xlsx'
output_file = '/Users/vagouz/Downloads/output.xlsx'
def update_execl_format(excel_file):
    # Load the workbook and select the active worksheet
    workbook = load_workbook(excel_file)
    sheet = workbook.active

    # Iterate over all cells in the worksheet and set wrapText to True
    for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrapText=True)

    # Save the changes to the workbook
    workbook.save(excel_file)

email_groups = {
    "audio": ["liechen@cisco.com","qunwan@cisco.com", "mingxuzh@cisco.com","borisl@cisco.com","huawan@cisco.com","feijie@cisco.com","mfei@cisco.com","bingyu2@cisco.com","mfei@cisco.com","yulyang@cisco.com","leiliu3@cisco.com"],
    "video": ["bojfan@cisco.com", "benzzhan@cisco.com","xiangzhc@cisco.com","xuexin@cisco.com","jiaying@cisco.com"],
    "sharing": ["shancxu@cisco.com","ebye@cisco.com","xiaolsun@cisco.com"],
    "transimit": ["jiqin@cisco.com","sasensar@cisco.com","smalimat@cisco.com","joycqu@cisco.com"],
    "mics": ["junga@cisco.com"]
}

def find_email_group(email):
    # Iterate over the groups to find the email
    for group, emails in email_groups.items():
        if email in emails:
            return group
    return "Other"

def get_api_response(api_url, token):
    # ciscospark://us/ROOM/3b5d69c0-8053-11ea-9696-51434787666c webexteams://im?space=3b5d69c0-8053-11ea-9696-51434787666c
    # Define the headers with the token
    headers = {
        'Authorization': f'Bearer {token}'
    }

    try:
        # Make a GET request to the API
        response = requests.get(api_url, headers=headers)

        # Check if the request was successful
        if response.status_code == 200:
            # Parse the JSON response data
            response_data = response.json()
            # Print the response data
            # print("Response Data:", response_data)
            return response_data
        else:
            print(f"Failed to retrieve data. Status code: {response.status_code}")
            print("Response:", response.text)
            return None
    except requests.exceptions.RequestException as e:
        # Handle any exceptions that occur during the request
        print(f"An error occurred: {e}")
        return None

def get_data_from_cloud(total_count=100):
    # Example usage
    api_url = f"https://webexapis.com/v1/messages?roomId=Y2lzY29zcGFyazovL3VzL1JPT00vM2I1ZDY5YzAtODA1My0xMWVhLTk2OTYtNTE0MzQ3ODc2NjZj&max={total_count}"  # Replace with your API URL
    token = "OGY4M2Q0NTgtODQ1Yi00MTFhLWIwYzktNWZjY2EzMDYxNGJjODgyOWM0ZmEtMTZh_PF84_1eb65fdf-9643-417f-9974-ad72cae0e10f"          # Replace with your token
    response_data = get_api_response(api_url, token)
    # print("Response Data:", response_data)
    if response_data is None:
        return None
    
    # Convert the list of dictionaries to a DataFrame
    df_orignal = pd.DataFrame(response_data['items'])
    # Define the columns to keep
    columns_to_keep = ['text', 'personEmail', 'created']
    df_orignal = df_orignal[columns_to_keep]
    df_orignal = df_orignal.dropna(subset=['text'])
    df_orignal = df_orignal[df_orignal['text'].str.contains('EP request for', na=False)]
    df_orignal.to_excel(source_file, index=False)
    return df_orignal

def get_data_from_file():
    # Load the data from the CSV file
    df = pd.read_excel(source_file)
    return df


# Function to extract details from the 'text' column
def extract_details(row):
    text = row['text']
    
    # Extract details using regex
    jira_id = re.search(r'https://jira-eng-gpk2\.cisco\.com/jira/browse/(WEBEX|SPARK)-\d+', text)
    if jira_id is None:
        jira_id = re.search(r'\[(WEBEX|SPARK)-\d+\]', text)
    jira_id = jira_id.group(0).strip() if jira_id else None
    pr_link = re.search(r'https://sqbu-github\.cisco\.com/WebExSquared/wme/pull/\d+', text)
    if pr_link is None:
        pr_link = re.search(r'https://sqbu-github\.cisco\.com/WebexApps/webex-apps/pull/\d+', text)
    pr_link = pr_link.group(0).strip() if pr_link else None
    title = re.search(r'Title:(.*?)\n', text, re.DOTALL)
    title = title.group(1).strip() if title else None
    
    # Define markers and corresponding variables
    markers = {
        'Which of the below category does you': 'ep_category',
        'Explain why this can\'t wait until next GA': 'why_need_ep',
        'Is this a regression from current shipping product': 'is_regression',
        'How many users are impacted by this issue': 'issue_impact',
        'Explain why this is a low risk change': 'risk',
        'Which component(s) and clients type are impacted': 'issue_scope',
        'How do you plan to avoid this issue from happening again late in our release process': 'what_improve'
    }
    
    # Split the text into lines
    lines = text.splitlines()
    
    # Iterate over the lines and extract details
    current_marker = None
    current_value = ''
    marker_values = {}
    
    for line in lines:
        # Check if the line contains any marker
        contains_marker = False
        for marker, variable in markers.items():
            if marker in line:
                # Save the previous marker's value
                if current_marker:
                    marker_values[markers[current_marker]] = current_value.strip()
                # Set the new marker
                current_marker = marker
                current_value = ''
                contains_marker = True
                break
        if contains_marker:
            continue
        # If no marker is found, accumulate the value
        if current_marker:
            current_value += line
    
    # Save the last marker's value
    if current_marker:
        marker_values[markers[current_marker]] = current_value.strip()
    
    #
    wme_module = find_email_group(row['personEmail'])

    # Format the date
    dt = datetime.strptime(row['created'], "%Y-%m-%dT%H:%M:%S.%fZ")
    formatted_date = dt.strftime("%Y-%m")

    # Normalize the values
    if 'ep_category' in marker_values:
        if marker_values['ep_category'] is None:
            marker_values['ep_category'] = 'Other'
        if marker_values['ep_category'].lower().startswith('bug') or 'bug' in marker_values['ep_category'].lower():
            marker_values['ep_category'] = 'Bug'
        else:
            marker_values['ep_category'] = 'Other'
    
    if 'is_regression' in marker_values:
        if marker_values['is_regression'].lower().startswith('yes') or 'regression' in marker_values['is_regression'].lower():
            marker_values['is_regression'] = 'Yes'
        else:
            marker_values['is_regression'] = 'No'

    return pd.Series({
        'JiraID': jira_id,
        'PRLink': pr_link,
        'Title': title,
        'Month': formatted_date,
        'wme_module': wme_module,
        'EP-Category': marker_values['ep_category'] if 'ep_category' in marker_values else None,
        'WhyNeedEP': marker_values['why_need_ep'] if 'why_need_ep' in marker_values else None,
        'IsRegression': marker_values['is_regression']  if 'is_regression' in marker_values else None,
        'IssueImpact': marker_values['issue_impact'] if 'issue_impact' in marker_values else None,
        'Risk': marker_values['risk']  if 'risk' in marker_values else None,
        'IssueScope': marker_values['issue_scope'] if 'issue_scope' in marker_values else None,
        'WhatImprove': marker_values['what_improve'] if 'what_improve' in marker_values else None
    })

fetch_data = False
if fetch_data:
    df_orignal = get_data_from_cloud(1000)
    if df_orignal is None:
        print('Failed to retrieve data from cloud')
        exit()

df_update = get_data_from_file() 
if df_update is None:
    print('Failed to retrieve data from file')
    exit()

# Apply the function to each row to add new columns
df_update = df_update.join(df_update.apply(extract_details, axis=1))
# re-order columns
df_update = df_update[['text', 'wme_module', 'Month', 'IsRegression', 'EP-Category', 'JiraID', 'PRLink', 'Title',  'WhyNeedEP',  'IssueImpact', 'Risk', 'IssueScope', 'WhatImprove',  'personEmail', 'created']]
# Save the updated DataFrame to a new Excel file
df_update.to_excel(output_file, index=False)
update_execl_format(output_file)